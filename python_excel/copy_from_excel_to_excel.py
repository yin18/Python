# write a program in python to vopy data from data9 file to new file called ""
import openpyxl
import xlsxwriter

data9_copied = xlsxwriter.Workbook("data9_copied.xlsx")
data9_copied.close()

# opening source file
path = "data9.xlsx"
data9_workbook = openpyxl.load_workbook(path)
data9_worksheet = data9_workbook.worksheets[0]

# opening the destination excel file
new_path = "data9_copied.xlsx"
data9_copied_workbook = openpyxl.load_workbook(new_path)
data9_copied_worksheet = data9_workbook.active

mc = data9_copied_worksheet.max_column
mr = data9_copied_worksheet.max_row

for x in range(1, mr + 1):
    for y in range(1, mc + 1):
        cell_content = data9_worksheet.cell(row=x, column=y)

        data9_copied_worksheet.cell(row=x, column=y).value = cell_content.value
