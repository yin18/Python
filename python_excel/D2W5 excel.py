import xlsxwriter

data9 = xlsxwriter.Workbook("data9.xlsx")

worksheet = data9.add_worksheet("data9.xlsx")

#rows and columns are zero indexed
#first cell in worksheet A1 is (0,0), B1 is (0,1)
worksheet.write("A1", "data 9")
worksheet.write("A2", "Isabella")
worksheet.write("A3", "Shahrukh")
worksheet.write("A4", "Zoe")

row = 0
column = 1


info = ["Name", "Sassy", "CJ"]
for name in info:
    worksheet.write(row, column, name)
    row +=1
data9.close()
print(data9)



import openpyxl

#give the location of our file
path = "data9.xlsx"

#creating an object
data9_object = openpyxl.load_workbook(path)

#storing the data from active sheet into am pbject called data9
data9_object = data9_object.active #editable
data9_sheet = data9_object.title
print(data9_sheet) #title of the sheet

#retrieves data in cell A1
get_data = data9_object.cell(row=1, column=1)
print(get_data.value)

data9_object.cell(row=2, column=1).value = "changed"
print(data9_object.cell(row=2, column=1).value)

#sets the title of the sheet
data9_object.title = "Trainees"
print(data9_object.title)

#number of rows and columns
total_row = data9_object.max_row
total_col = data9_object.max_column

print("total row: ", total_row)
print("total column: ", total_col)











